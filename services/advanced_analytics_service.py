"""Сервис расширенной аналитики с метриками для визуализации."""

from typing import Dict, List, Optional, Tuple, Union
from datetime import datetime, timedelta
from dataclasses import dataclass
from enum import Enum
from database.connection import get_db_pool


@dataclass
class ConversionMetrics:
    """Метрики конверсии."""
    total_registrations: int
    approved: int
    rejected: int
    pending: int
    conversion_rate: float
    approval_rate: float
    rejection_rate: float


@dataclass
class RetentionMetrics:
    """Метрики удержания."""
    total_users: int
    returning_users: int
    new_users: int
    retention_rate: float
    churn_rate: float


@dataclass
class TimeSeriesPoint:
    """Точка временного ряда."""
    timestamp: datetime
    value: float
    label: str


class MetricPeriod(Enum):
    """Периоды для метрик."""
    HOURLY = "hourly"
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"


class AdvancedAnalyticsService:
    """Сервис расширенной аналитики."""
    
    async def get_conversion_funnel(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Dict:
        """
        Получает воронку конверсии с процентами.
        
        Этапы воронки:
        1. Начали регистрацию
        2. Заполнили имя
        3. Заполнили телефон
        4. Заполнили карту
        5. Загрузили фото
        6. Отправили на модерацию
        7. Одобрено
        
        Returns:
            Данные для визуализации воронки
        """
        pool = get_db_pool()
        
        if not start_date:
            start_date = datetime.now() - timedelta(days=30)
        if not end_date:
            end_date = datetime.now()
        
        async with pool.connection() as conn:
            # Общее количество начавших регистрацию
            cursor = await conn.execute(
                """
                SELECT COUNT(DISTINCT user_id) 
                FROM analytics_events 
                WHERE event_type = 'registration_started'
                AND created_at BETWEEN ? AND ?
                """,
                (start_date, end_date)
            )
            started = (await cursor.fetchone())[0] or 1  # Избегаем деления на 0
            
            # Завершили регистрацию
            cursor = await conn.execute(
                """
                SELECT COUNT(DISTINCT user_id)
                FROM analytics_events
                WHERE event_type = 'registration_completed'
                AND created_at BETWEEN ? AND ?
                """,
                (start_date, end_date)
            )
            completed = (await cursor.fetchone())[0]
            
            # Одобрено
            cursor = await conn.execute(
                """
                SELECT COUNT(*)
                FROM participants
                WHERE status = 'approved'
                AND registration_date BETWEEN ? AND ?
                """,
                (start_date, end_date)
            )
            approved = (await cursor.fetchone())[0]
        
        # Вычисляем проценты
        funnel_stages = [
            {
                "stage": "Начали регистрацию",
                "count": started,
                "percentage": 100.0,
                "drop_off": 0
            },
            {
                "stage": "Завершили регистрацию",
                "count": completed,
                "percentage": round((completed / started) * 100, 2),
                "drop_off": round(((started - completed) / started) * 100, 2)
            },
            {
                "stage": "Одобрено",
                "count": approved,
                "percentage": round((approved / started) * 100, 2),
                "drop_off": round(((completed - approved) / completed) * 100, 2) if completed > 0 else 0
            }
        ]
        
        return {
            "stages": funnel_stages,
            "overall_conversion": round((approved / started) * 100, 2),
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            }
        }
    
    async def get_conversion_metrics(
        self,
        period_days: int = 30
    ) -> ConversionMetrics:
        """
        Получает метрики конверсии за период.
        
        Args:
            period_days: Количество дней для анализа
            
        Returns:
            Метрики конверсии
        """
        pool = get_db_pool()
        start_date = datetime.now() - timedelta(days=period_days)
        
        async with pool.connection() as conn:
            cursor = await conn.execute(
                """
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN status = 'approved' THEN 1 ELSE 0 END) as approved,
                    SUM(CASE WHEN status = 'rejected' THEN 1 ELSE 0 END) as rejected,
                    SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending
                FROM participants
                WHERE registration_date >= ?
                """,
                (start_date,)
            )
            row = await cursor.fetchone()
            
            total = row[0] or 1
            approved = row[1] or 0
            rejected = row[2] or 0
            pending = row[3] or 0
            
            return ConversionMetrics(
                total_registrations=total,
                approved=approved,
                rejected=rejected,
                pending=pending,
                conversion_rate=round((approved / total) * 100, 2),
                approval_rate=round((approved / (total - pending)) * 100, 2) if (total - pending) > 0 else 0,
                rejection_rate=round((rejected / (total - pending)) * 100, 2) if (total - pending) > 0 else 0
            )
    
    async def get_retention_metrics(
        self,
        period_days: int = 30
    ) -> RetentionMetrics:
        """
        Получает метрики удержания (retention) пользователей.
        
        Args:
            period_days: Количество дней для анализа
            
        Returns:
            Метрики retention и churn
        """
        pool = get_db_pool()
        start_date = datetime.now() - timedelta(days=period_days)
        previous_period_start = start_date - timedelta(days=period_days)
        
        async with pool.connection() as conn:
            # Пользователи текущего периода
            cursor = await conn.execute(
                """
                SELECT COUNT(DISTINCT telegram_id)
                FROM participants
                WHERE registration_date >= ?
                """,
                (start_date,)
            )
            current_users = (await cursor.fetchone())[0] or 0
            
            # Пользователи предыдущего периода
            cursor = await conn.execute(
                """
                SELECT COUNT(DISTINCT telegram_id)
                FROM participants
                WHERE registration_date BETWEEN ? AND ?
                """,
                (previous_period_start, start_date)
            )
            previous_users = (await cursor.fetchone())[0] or 1
            
            # Вернувшиеся пользователи (были в предыдущем и есть в текущем)
            cursor = await conn.execute(
                """
                SELECT COUNT(DISTINCT p1.telegram_id)
                FROM participants p1
                WHERE p1.registration_date BETWEEN ? AND ?
                AND EXISTS (
                    SELECT 1 FROM participants p2
                    WHERE p2.telegram_id = p1.telegram_id
                    AND p2.registration_date >= ?
                )
                """,
                (previous_period_start, start_date, start_date)
            )
            returning = (await cursor.fetchone())[0] or 0
            
            new_users = current_users - returning
            retention_rate = round((returning / previous_users) * 100, 2) if previous_users > 0 else 0
            churn_rate = round(100 - retention_rate, 2)
            
            return RetentionMetrics(
                total_users=current_users,
                returning_users=returning,
                new_users=new_users,
                retention_rate=retention_rate,
                churn_rate=churn_rate
            )
    
    async def get_time_series(
        self,
        metric: str,
        period: MetricPeriod = MetricPeriod.DAILY,
        days: int = 30
    ) -> List[TimeSeriesPoint]:
        """
        Получает временной ряд для метрики.
        
        Args:
            metric: Название метрики (registrations, approvals, rejections)
            period: Период группировки (hourly, daily, weekly)
            days: Количество дней для анализа
            
        Returns:
            Список точек временного ряда
        """
        pool = get_db_pool()
        start_date = datetime.now() - timedelta(days=days)
        
        # Определяем формат группировки
        if period == MetricPeriod.HOURLY:
            date_format = "%Y-%m-%d %H:00:00"
        elif period == MetricPeriod.DAILY:
            date_format = "%Y-%m-%d"
        elif period == MetricPeriod.WEEKLY:
            date_format = "%Y-W%W"
        else:  # MONTHLY
            date_format = "%Y-%m"
        
        async with pool.connection() as conn:
            if metric == "registrations":
                query = f"""
                    SELECT 
                        strftime('{date_format}', registration_date) as period,
                        COUNT(*) as value
                    FROM participants
                    WHERE registration_date >= ?
                    GROUP BY period
                    ORDER BY period
                """
            elif metric == "approvals":
                query = f"""
                    SELECT 
                        strftime('{date_format}', registration_date) as period,
                        COUNT(*) as value
                    FROM participants
                    WHERE status = 'approved' AND registration_date >= ?
                    GROUP BY period
                    ORDER BY period
                """
            elif metric == "rejections":
                query = f"""
                    SELECT 
                        strftime('{date_format}', registration_date) as period,
                        COUNT(*) as value
                    FROM participants
                    WHERE status = 'rejected' AND registration_date >= ?
                    GROUP BY period
                    ORDER BY period
                """
            else:
                return []
            
            cursor = await conn.execute(query, (start_date,))
            rows = await cursor.fetchall()
            
            time_series = []
            for row in rows:
                time_series.append(TimeSeriesPoint(
                    timestamp=datetime.strptime(row[0], date_format.replace('%W', '01')),
                    value=float(row[1]),
                    label=row[0]
                ))
            
            return time_series
    
    async def get_activity_heatmap(
        self,
        days: int = 30
    ) -> Dict[str, Dict[int, int]]:
        """
        Создает heatmap активности по дням недели и часам.
        
        Returns:
            Словарь {день_недели: {час: количество}}
        """
        pool = get_db_pool()
        start_date = datetime.now() - timedelta(days=days)
        
        async with pool.connection() as conn:
            cursor = await conn.execute(
                """
                SELECT 
                    CAST(strftime('%w', registration_date) AS INTEGER) as day_of_week,
                    CAST(strftime('%H', registration_date) AS INTEGER) as hour_of_day,
                    COUNT(*) as activity_count
                FROM participants
                WHERE registration_date >= ?
                GROUP BY day_of_week, hour_of_day
                ORDER BY day_of_week, hour_of_day
                """,
                (start_date,)
            )
            rows = await cursor.fetchall()
        
        # Инициализируем heatmap
        days_map = {
            0: "Воскресенье",
            1: "Понедельник",
            2: "Вторник",
            3: "Среда",
            4: "Четверг",
            5: "Пятница",
            6: "Суббота"
        }
        
        heatmap = {day_name: {hour: 0 for hour in range(24)} for day_name in days_map.values()}
        
        # Заполняем данными
        for row in rows:
            day_name = days_map[row[0]]
            hour = row[1]
            count = row[2]
            heatmap[day_name][hour] = count
        
        return heatmap
    
    async def get_cohort_analysis(
        self,
        cohort_size_weeks: int = 1
    ) -> List[Dict]:
        """
        Cohort анализ - как ведут себя пользователи, зарегистрированные в одно время.
        
        Args:
            cohort_size_weeks: Размер когорты в неделях
            
        Returns:
            Данные для cohort analysis
        """
        pool = get_db_pool()
        
        async with pool.connection() as conn:
            cursor = await conn.execute(
                """
                SELECT 
                    strftime('%Y-W%W', registration_date) as cohort_week,
                    COUNT(*) as cohort_size,
                    SUM(CASE WHEN status = 'approved' THEN 1 ELSE 0 END) as approved_count,
                    SUM(CASE WHEN status = 'rejected' THEN 1 ELSE 0 END) as rejected_count
                FROM participants
                WHERE registration_date >= date('now', '-12 weeks')
                GROUP BY cohort_week
                ORDER BY cohort_week
                """
            )
            rows = await cursor.fetchall()
        
        cohorts = []
        for row in rows:
            cohort = {
                "cohort": row[0],
                "size": row[1],
                "approved": row[2],
                "rejected": row[3],
                "approval_rate": round((row[2] / row[1]) * 100, 2) if row[1] > 0 else 0
            }
            cohorts.append(cohort)
        
        return cohorts
    
    async def get_real_time_stats(self) -> Dict:
        """
        Получает статистику в реальном времени для dashboard.
        
        Returns:
            Текущие показатели системы
        """
        pool = get_db_pool()
        now = datetime.now()
        today_start = datetime(now.year, now.month, now.day)
        
        async with pool.connection() as conn:
            # Статистика за сегодня
            cursor = await conn.execute(
                """
                SELECT 
                    COUNT(*) as today_total,
                    SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as today_pending,
                    SUM(CASE WHEN status = 'approved' THEN 1 ELSE 0 END) as today_approved,
                    SUM(CASE WHEN status = 'rejected' THEN 1 ELSE 0 END) as today_rejected
                FROM participants
                WHERE registration_date >= ?
                """,
                (today_start,)
            )
            today_stats = await cursor.fetchone()
            
            # Статистика за последний час
            hour_ago = now - timedelta(hours=1)
            cursor = await conn.execute(
                """
                SELECT COUNT(*) 
                FROM participants
                WHERE registration_date >= ?
                """,
                (hour_ago,)
            )
            last_hour = (await cursor.fetchone())[0]
            
            # Общая статистика
            cursor = await conn.execute(
                """
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending,
                    SUM(CASE WHEN status = 'approved' THEN 1 ELSE 0 END) as approved,
                    SUM(CASE WHEN status = 'rejected' THEN 1 ELSE 0 END) as rejected
                FROM participants
                """
            )
            total_stats = await cursor.fetchone()
        
        return {
            "today": {
                "total": today_stats[0] or 0,
                "pending": today_stats[1] or 0,
                "approved": today_stats[2] or 0,
                "rejected": today_stats[3] or 0
            },
            "last_hour": last_hour or 0,
            "all_time": {
                "total": total_stats[0] or 0,
                "pending": total_stats[1] or 0,
                "approved": total_stats[2] or 0,
                "rejected": total_stats[3] or 0
            },
            "timestamp": now.isoformat()
        }
    
    async def get_history_window(self) -> Tuple[Optional[datetime], Optional[datetime]]:
        """Получить минимальную и максимальную дату регистрации для определения полного периода."""
        pool = get_db_pool()
        async with pool.connection() as conn:
            cursor = await conn.execute(
                "SELECT MIN(registration_date), MAX(registration_date) FROM participants"
            )
            row = await cursor.fetchone()
            min_date = row[0]
            max_date = row[1]
            return min_date, max_date
    
    async def export_analytics_report(
        self,
        format: str = "xlsx",
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ):
        """
        Экспортирует полный аналитический отчет.
        
        Args:
            format: Формат экспорта (json, xlsx)
            start_date: Начальная дата (если None, берется минимальная дата из БД)
            end_date: Конечная дата (если None, берется максимальная дата из БД)
            
        Returns:
            Полный отчет со всеми метриками (Dict для JSON, BytesIO для XLSX)
        """
        # Если даты не указаны, получаем полный период из БД
        if not start_date or not end_date:
            min_date, max_date = await self.get_history_window()
            if not start_date:
                start_date = min_date or (datetime.now() - timedelta(days=30))
            if not end_date:
                end_date = max_date or datetime.now()
        
        # Вычисляем количество дней для метрик
        period_days = (end_date - start_date).days
        if period_days <= 0:
            period_days = 30
        
        # Собираем все метрики за весь период
        conversion = await self.get_conversion_metrics(period_days)
        retention = await self.get_retention_metrics(period_days)
        funnel = await self.get_conversion_funnel(start_date, end_date)
        time_series_reg = await self.get_time_series("registrations", MetricPeriod.DAILY, period_days)
        heatmap = await self.get_activity_heatmap(period_days)
        cohorts = await self.get_cohort_analysis()
        real_time = await self.get_real_time_stats()
        
        if format == "xlsx":
            # Экспорт в Excel
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.formatting.rule import ColorScaleRule
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            wb = Workbook()
            
            # Стили
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Лист "Summary"
            ws = wb.active
            ws.title = "Сводка"
            ws['A1'] = "Расширенный аналитический отчет"
            ws['A1'].font = title_font
            ws.merge_cells('A1:B1')
            
            ws['A3'] = "Период"
            ws['B3'] = f"{start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}"
            ws['A3'].font = Font(bold=True)
            ws['B3'].font = Font(bold=True)
            
            row = 5
            ws[f'A{row}'] = "Метрика"
            ws[f'B{row}'] = "Значение"
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].font = header_font
            
            summary_data = [
                ("Всего регистраций", conversion.total_registrations),
                ("Одобрено", conversion.approved),
                ("Отклонено", conversion.rejected),
                ("На модерации", conversion.pending),
                ("Conversion Rate", f"{conversion.conversion_rate}%"),
                ("Approval Rate", f"{conversion.approval_rate}%"),
                ("Rejection Rate", f"{conversion.rejection_rate}%"),
                ("Всего пользователей", retention.total_users),
                ("Новые пользователи", retention.new_users),
                ("Вернувшиеся", retention.returning_users),
                ("Retention Rate", f"{retention.retention_rate}%"),
                ("Churn Rate", f"{retention.churn_rate}%"),
            ]
            
            row = 6
            for metric, value in summary_data:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = value
                row += 1
            
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            
            # Лист "Воронка"
            ws2 = wb.create_sheet("Воронка")
            ws2['A1'] = "Этап"
            ws2['B1'] = "Количество"
            ws2['C1'] = "Процент"
            ws2['D1'] = "Отсев"
            
            for col in ['A', 'B', 'C', 'D']:
                ws2[f'{col}1'].fill = header_fill
                ws2[f'{col}1'].font = header_font
                ws2[f'{col}1'].border = border
            
            row = 2
            for stage in funnel.get("stages", []):
                ws2[f'A{row}'] = stage.get("stage", "")
                ws2[f'B{row}'] = stage.get("count", 0)
                ws2[f'C{row}'] = f"{stage.get('percentage', 0)}%"
                ws2[f'D{row}'] = f"{stage.get('drop_off', 0)}%"
                for col in ['A', 'B', 'C', 'D']:
                    ws2[f'{col}{row}'].border = border
                row += 1
            
            for col in ['A', 'B', 'C', 'D']:
                ws2.column_dimensions[col].width = 20
            
            # Лист "Временной ряд"
            ws3 = wb.create_sheet("Временной ряд")
            ws3['A1'] = "Дата"
            ws3['B1'] = "Регистрации"
            
            for col in ['A', 'B']:
                ws3[f'{col}1'].fill = header_fill
                ws3[f'{col}1'].font = header_font
                ws3[f'{col}1'].border = border
            
            row = 2
            for point in time_series_reg:
                ws3[f'A{row}'] = point.label
                ws3[f'B{row}'] = point.value
                for col in ['A', 'B']:
                    ws3[f'{col}{row}'].border = border
                row += 1
            
            ws3.column_dimensions['A'].width = 20
            ws3.column_dimensions['B'].width = 15
            
            # Лист "Тепловая карта"
            ws4 = wb.create_sheet("Тепловая карта")
            ws4['A1'] = "День"
            for hour in range(0, 24, 3):
                col_letter = get_column_letter(2 + hour // 3)
                ws4[f'{col_letter}1'] = f"{hour:02d}:00"
                ws4[f'{col_letter}1'].fill = header_fill
                ws4[f'{col_letter}1'].font = header_font
                ws4[f'{col_letter}1'].border = border
            
            ws4['A1'].fill = header_fill
            ws4['A1'].font = header_font
            ws4['A1'].border = border
            
            days_order = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
            row = 2
            for day_name in days_order:
                if day_name in heatmap:
                    ws4[f'A{row}'] = day_name
                    ws4[f'A{row}'].border = border
                    col_idx = 2
                    for hour in range(0, 24, 3):
                        col_letter = get_column_letter(col_idx)
                        count = heatmap[day_name].get(hour, 0)
                        ws4[f'{col_letter}{row}'] = count
                        ws4[f'{col_letter}{row}'].border = border
                        col_idx += 1
                    row += 1
            
            ws4.column_dimensions['A'].width = 15
            for hour in range(0, 24, 3):
                col_letter = get_column_letter(2 + hour // 3)
                ws4.column_dimensions[col_letter].width = 12
            
            # Лист "Когорты"
            ws5 = wb.create_sheet("Когорты")
            ws5['A1'] = "Когорта"
            ws5['B1'] = "Размер"
            ws5['C1'] = "Одобрено"
            ws5['D1'] = "Отклонено"
            ws5['E1'] = "Approval Rate"
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws5[f'{col}1'].fill = header_fill
                ws5[f'{col}1'].font = header_font
                ws5[f'{col}1'].border = border
            
            row = 2
            for cohort in cohorts:
                ws5[f'A{row}'] = cohort.get("cohort", "")
                ws5[f'B{row}'] = cohort.get("size", 0)
                ws5[f'C{row}'] = cohort.get("approved", 0)
                ws5[f'D{row}'] = cohort.get("rejected", 0)
                ws5[f'E{row}'] = f"{cohort.get('approval_rate', 0)}%"
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws5[f'{col}{row}'].border = border
                row += 1
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws5.column_dimensions[col].width = 18
            
            # Сохраняем в BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        
        # JSON формат (для обратной совместимости)
        report = {
            "generated_at": datetime.now().isoformat(),
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "conversion_metrics": {
                "total_registrations": conversion.total_registrations,
                "approved": conversion.approved,
                "rejected": conversion.rejected,
                "pending": conversion.pending,
                "conversion_rate": conversion.conversion_rate,
                "approval_rate": conversion.approval_rate,
                "rejection_rate": conversion.rejection_rate
            },
            "retention_metrics": {
                "total_users": retention.total_users,
                "returning_users": retention.returning_users,
                "new_users": retention.new_users,
                "retention_rate": retention.retention_rate,
                "churn_rate": retention.churn_rate
            },
            "funnel": funnel,
            "time_series": {
                "registrations": [
                    {"date": point.label, "value": point.value}
                    for point in time_series_reg
                ]
            },
            "activity_heatmap": heatmap,
            "cohort_analysis": cohorts,
            "real_time_stats": real_time,
            "format": format
        }
        
        return report


# Глобальный экземпляр
advanced_analytics_service = AdvancedAnalyticsService()

